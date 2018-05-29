# 从xlsx读取输出到sqlite数据库
import base64
import csv
import os
import re
import sqlite3

from openpyxl import load_workbook

xlsxname = 'news.xlsx'
datebasename = 'data.db'


def readxlsx(filename):
    wb = load_workbook(filename)
    ws = wb['新闻条目'.decode('utf-8')]
    row = ws.iter_rows(min_row=2)
    return row


def createTable(datebasename):
    conn = sqlite3.connect(datebasename)
    cursor = conn.cursor()
    cursor.execute(
        'create table news (id interger,title varchar(255),content text,clicks int,date text)')
    cursor.close()
    conn.commit()
    conn.close()


def addData(row):
    conn = sqlite3.connect('data.db')
    conn.text_factory = str
    cursor = conn.cursor()
    dr = re.compile(r'<br\s*?/?>')
    text = row[2].value
    missing_padding = 4 - len(text) % 4
    if missing_padding:
        text += b'=' * missing_padding
    try:
        content = base64.b64decode(text)
    except TypeError:
        print 'error'
        content = 'error'
    # with open("news.csv","w") as f:
    #     writer = csv.writer(f)
    #     writer.writerow([row[0].value, row[1].value, dr.sub('', content), row[3].value, row[4].value])
    cursor.execute('insert into news values (?,?,?,?,?)', (row[0].value, row[1].value, dr.sub(
        '', content), row[4].value, row[3].value))
    cursor.close()
    conn.commit()
    conn.close()


rows = readxlsx(xlsxname)
if not os.path.exists(datebasename):
    createTable(datebasename)
for row in rows:
    addData(row)
